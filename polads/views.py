from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from .forms import AddAdicional, AddTurno
from administracion.models import Comisaria
from .models import Adicional, Turno
from cores.models import Cores
from .filters import RendicionFilter
from users.models import Profile
from datetime import date, timedelta, datetime
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.db.models import F, Sum, ExpressionWrapper, fields, Case, When, Value, CharField
from mensajes.models import Mensaje
from django.utils.safestring import mark_safe
from .utils import CalendarioTurnos

# agregar adicional:

@login_required
def nuevo_adicional(request):
    form = AddAdicional()
    if request.method == 'POST':
        form = AddAdicional(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.encargado = request.user
            instance.puede_rendir = True
            instance.save()
            messages.success(request, f'{request.user} creaste un adicional exitosamente.')
            return redirect('mis-adicionales')
        else:
            print(form.errors.items())
    context = {'form': form}
    return render(request, 'polads/nuevo_adicional.html', context)

# eliminar turno:

@login_required
def deleteturno(request, field):
    previous_page = request.META['HTTP_REFERER']
    turnotodelete = Turno.objects.get(id=field)
    turnotodelete.delete()
    return redirect(previous_page)

# listado de adicionales habilitados:

@login_required
def mis_adicionales(request):
    today = datetime.now()
    usuario = request.user.profile.tipo_usuario
    dependencia = request.user.profile.administrador_de
    if usuario == 'Encargado':
        adicionales = Adicional.objects.filter(encargado=request.user) # sólo muestra aquellos adicionales creados por el encargado
    else:
        adicionales = Adicional.objects.filter(encargado=request.user)|Adicional.objects.filter(comisaria=dependencia) # muestra los adicionales de la comisaría
    context = {'adicionales': adicionales, 'today': today}
    return render(request, 'polads/mis_adicionales.html', context)

# agregar turno:

@login_required
def agregar_turnos(request, pk):
    adicional = Adicional.objects.get(pk=pk)
    turnos = Turno.objects.filter(polad=adicional).order_by('-pk')[:300]
    errors = ''
    form = AddTurno()
    if request.method == 'POST':
        form = AddTurno(request.POST)
        if form.is_valid():

            instance = form.save(commit=False)

            # validación de turnos:

            all_turnos = Turno.objects.filter(efectivo=instance.efectivo)
            all_cores = Cores.objects.filter(efectivo=instance.efectivo)

            cant_cores = Cores.objects.filter(efectivo=instance.efectivo, ingreso__month=instance.ingreso.month, ingreso__year=instance.ingreso.year).aggregate(
            horas=Sum(ExpressionWrapper(F('egreso') - F('ingreso'), output_field=fields.DurationField())))

            cant_polad = Turno.objects.filter(efectivo=instance.efectivo, ingreso__month=instance.ingreso.month, ingreso__year=instance.ingreso.year).aggregate(
            horas=Sum(ExpressionWrapper(F('egreso') - F('ingreso'), output_field=fields.DurationField())))

            try:
                cant_cores = int(cant_cores.get('horas').total_seconds()/60/60)
            except:
                cant_cores = 0

            try:
                cant_polad = int(cant_polad.get('horas').total_seconds()/60/60)
            except:
                cant_polad = 0

            horas_mensuales = cant_cores + cant_polad

            es_valido = True

            for each in all_turnos:
                if instance.ingreso < each.egreso and each.ingreso < instance.egreso:
                        errors = 'Turno inválido para el efectivo ' + str(instance.efectivo) + ". Comunicate con " + str(each.polad.encargado) + "."
                        try:
                            administrador = Profile.objects.get(administrador_de=each.polad.comisaria)
                            mensaje = Mensaje(remitente=request.user, destinatario=administrador.user, mensaje=f'Mensaje automático. El usuario {request.user} ha intentado cargar un turno inválido de Adicional para el efectivo {instance.efectivo} en el servicio {adicional.nombre} el día {instance.ingreso.day}/{instance.ingreso.month} a partir de las {instance.ingreso.hour}.')
                            mensaje.save()
                        except Profile.DoesNotExist:
                            administrador = None
                        es_valido = False

            for each in all_cores:
                if instance.ingreso < each.egreso and each.ingreso < instance.egreso:
                        responsable = Profile.objects.get(administrador_de=each.dependencia)
                        errors = 'Turno inválido para el efectivo ' + str(instance.efectivo) + ". Comunicate con " + str(responsable.user) + "."
                        mensaje = Mensaje(remitente=request.user, destinatario=responsable.user, mensaje=f'Mensaje automático. El usuario {request.user} ha intentado crear un turno inválido de Adicional para el efectivo {instance.efectivo} en el servicio {adicional.nombre} el día {instance.ingreso.day}/{instance.ingreso.month} a partir de las {instance.ingreso.hour}.')
                        mensaje.save()
                        es_valido = False

            cant_instance = int((instance.egreso - instance.ingreso).seconds/3600)

            if instance.ingreso > instance.egreso:
                errors = "El ingreso no puede ser posterior al egreso."
                es_valido = False

            if cant_instance + cant_polad > 180:
                errors = "El efectivo ha superado el límite de horas para hacer Adicionales."
                es_valido = False

            # procesar turnos y horarios:

            if es_valido == True:

                if instance.ingreso.day == instance.egreso.day:
                    instance_valid = Turno(ingreso=instance.ingreso, egreso=instance.egreso, efectivo=instance.efectivo, polad=adicional)
                    instance_valid.save()

                elif instance.ingreso.day != instance.egreso.day:

                    if instance.egreso.month - instance.ingreso.month == 1 or instance.egreso.year - instance.ingreso.year == 1:
                        instance_valid_1 = Turno(ingreso=instance.ingreso, egreso=instance.egreso.replace(hour=00), efectivo=instance.efectivo, polad=adicional)
                        instance_valid_1.save()
                        instance_valid_2 = Turno(ingreso=instance.egreso.replace(hour=00), egreso=instance.egreso, efectivo=instance.efectivo, polad=adicional)
                        if instance_valid_2.egreso.hour == 00: pass
                        else: instance_valid_2.save()

                    if instance.egreso.day - instance.ingreso.day == 1:
                       instance_valid_1 = Turno(ingreso=instance.ingreso, egreso=instance.egreso.replace(hour=00), efectivo=instance.efectivo, polad=adicional)
                       instance_valid_1.save()
                       instance_valid_2 = Turno(ingreso=instance.egreso.replace(hour=00), egreso=instance.egreso, efectivo=instance.efectivo, polad=adicional)
                       if instance_valid_2.egreso.hour == 00: pass
                       else: instance_valid_2.save()

                    if instance.egreso.day - instance.ingreso.day > 1:
                        delta = instance.egreso - instance.ingreso
                        instance_valid_1 = Turno(ingreso=instance.ingreso, egreso=instance.ingreso.replace(day=instance.ingreso.day + 1, hour=00), efectivo=instance.efectivo, polad=adicional)
                        instance_valid_1.save()
                        for i in range(delta.days):
                            day = instance.ingreso + timedelta(days=i)
                            instance_valid_2 = Turno(ingreso=day.replace(day=day.day + 1, hour=00), egreso=day.replace(day=day.day + 2, hour=00), efectivo=instance.efectivo, polad=adicional)
                            if instance_valid_2.ingreso.day == instance.egreso.day: pass
                            else: instance_valid_2.save()
                        instance_valid_3 = Turno(ingreso=instance.egreso.replace(hour=00), egreso=instance.egreso, efectivo=instance.efectivo, polad=adicional)
                        instance_valid_3.save()

        else:
            print(form.errors.items())

    context = {'adicional': adicional, 'form':form, 'turnos': turnos, 'errors': errors}
    return render(request, 'polads/agregar_turnos.html', context)

# vista de calendario

@login_required
def calendar(request, pk, year, month):
    adicional = Adicional.objects.get(pk=pk)
    turnos = Turno.objects.filter(polad=adicional, ingreso__year=year, ingreso__month=month).order_by('ingreso')
    locale = 'es_ES'
    cal = CalendarioTurnos(turnos, locale).formatmonth(year, month)
    context = {'calendar': mark_safe(cal), 'adicional': adicional}
    return render(request, 'polads/calendario.html', context)

# filtrado de turnos para rendir:

@login_required
def rendir_adicional(request, pk):

    adicional = Adicional.objects.get(pk=pk)
    turnos = Turno.objects.filter(polad=adicional).order_by('-pk')
    myFilter = RendicionFilter(request.GET, queryset=turnos)
    turnos = myFilter.qs

    turnos_filtrados = [turno.id for turno in turnos]
    request.session['export_querset'] = turnos_filtrados

    context = {'adicional': adicional, 'turnos': turnos, 'myFilter': myFilter}
    return render(request, 'polads/rendir_adicional.html', context)

# validación previa de turnos a generar:

@login_required
def exportar(request, pk):

    adicional = Adicional.objects.get(pk=pk)
    turnos = [Turno.objects.get(id=id) for id in request.session['export_querset']]

    horas_totales = 0
    for turno in turnos:
        horas_totales = horas_totales + turno.get_diff

    cant = len(turnos)

    valor_100 = 0
    if adicional.categoria == "1era" or adicional.categoria == "4ta":
        valor_100 = horas_totales * 210
    elif adicional.categoria == "2da" or adicional.categoria == "3era":
        valor_100 = horas_totales * 380
    else:
        valor_100 = horas_totales * 360

    mes = []
    anio = []
    for turno in turnos:
        mes.append(turno.ingreso.month)
        anio.append(turno.ingreso.year)
    if (len(set(mes)) == 1):
        if mes[0] == 1: show = "Enero de {}".format(anio[0])
        elif mes[0] == 2: show = "Febrero de {}".format(anio[0])
        elif mes[0] == 3: show = "Marzo de {}".format(anio[0])
        elif mes[0] == 4: show = "Abril de {}".format(anio[0])
        elif mes[0] == 5: show = "Mayo de {}".format(anio[0])
        elif mes[0] == 6: show = "Junio de {}".format(anio[0])
        elif mes[0] == 7: show = "Julio de {}".format(anio[0])
        elif mes[0] == 8: show = "Agosto de {}".format(anio[0])
        elif mes[0] == 9: show = "Septiembre de {}".format(anio[0])
        elif mes[0] == 10: show = "Octubre de {}".format(anio[0])
        elif mes[0] == 11: show = "Noviembre de {}".format(anio[0])
        elif mes[0] == 12: show = "Diciembre de {}".format(anio[0])
    else: show = "Parece que estás intentando rendir un adicional con turnos de meses distintos."

    request.session['export_month'] = show

    context = {'adicional': adicional, 'turnos': turnos, 'cant': cant, 'show': show, 'horas_totales': horas_totales, 'valor_100': valor_100}
    return render(request, 'polads/exportar.html', context)


# creación de excel:

@login_required
def descargar(request, pk):
    adicional = Adicional.objects.get(pk=pk)
    turnos = [Turno.objects.get(id=id) for id in request.session['export_querset']]

    # normalización de turnos y horarios

    a_cargar = []

    for turno in turnos:

        if turno.ingreso.day != turno.egreso.day:

            if turno.egreso.day - turno.ingreso.day == 1:
               turno1 = [turno.efectivo.efectivo_nombre, turno.ingreso.day, turno.ingreso.hour, 24, 24 - turno.ingreso.hour]
               turno2 = [turno.efectivo.efectivo_nombre, turno.egreso.day, 00, turno.egreso.hour, turno.egreso.hour - 00]
               a_cargar.append(turno1)
               if turno2[2] == turno2[3]: pass
               else: a_cargar.append(turno2)

            elif turno.egreso.month - turno.ingreso.month == 1:
                turno1 = [turno.efectivo.efectivo_nombre, turno.ingreso.day, turno.ingreso.hour, 24, 24 - turno.ingreso.hour]
                a_cargar.append(turno1)

            elif turno.egreso.day - turno.ingreso.day > 1:
               delta = turno.egreso - turno.ingreso
               turno1 = [turno.efectivo.efectivo_nombre, turno.ingreso.day, turno.ingreso.hour, 24, 24 - turno.ingreso.hour]
               a_cargar.append(turno1)
               for i in range(delta.days + 1):
                   day = turno.ingreso + timedelta(days=i)
                   turno2 = [turno.efectivo.efectivo_nombre, day.day, 00, 24, 24]
                   if day.day == turno.ingreso.day or day.day == turno.egreso.day: pass
                   else: a_cargar.append(turno2)
               turno3 = [turno.efectivo.efectivo_nombre, turno.egreso.day, 00, turno.egreso.hour, turno.egreso.hour - 00]
               a_cargar.append(turno3)

        else:
            turno = [turno.efectivo.efectivo_nombre, turno.ingreso.day,  turno.ingreso.hour, turno.egreso.hour, turno.egreso.hour - turno.ingreso.hour]
            a_cargar.append(turno)

    for a in a_cargar:
        for j in a_cargar:
            if j[0] == a[0] and j[1] == a[1] and j[2] == a[3]:  # si es el mismo efectivo, en el mismo dia, y el egreso de uno coincide con el ingreso siguiente:
                turno = [a[0], a[1], a[2], j[3], a[4] + j[4]]   # el turno sera: nombre, dia, ingreso 1, egreso 2, suma de los dos turnos
                a_cargar.append(turno)                          # adjunta el turno nuevo creado
                a_cargar.remove(a)                              # elimina el turno 1
                a_cargar.remove(j)                              # eliminar el turno 2

    a_cargar.sort(key = lambda a_cargar: a_cargar[0])           # ordena por efectivo

    for a in a_cargar:
        a[2] = str(a[2])
        a[3] = str(a[3])
        if len(a[2]) == 1:                                      # si el ingreso tiene un solo dígito
            a[2] = "0" + str(a[2])                              # agregar un 0, para que 2 se transforme en 02
        if len(a[3]) == 1:                                      # igual el egreso
            a[3] = "0" + str(a[3])

    final_list = []

    for a in a_cargar:
        a = [a[0], a[1], a[2]+" a "+a[3], a[4]]                 # convertir en cadena ingreso y egreso "02 a 06"
        final_list.append(a)

    # load excel

    rend = load_workbook('/administracion/files/Test.xlsx')

    # creación de planillas individuales

    for turno in turnos:
        pest = rend.sheetnames
        if turno.efectivo.efectivo_nombre not in pest:
            source = rend['Efectivo']
            target = rend.copy_worksheet(source)

            target.title = turno.efectivo.efectivo_nombre

            target['A8'] = turno.efectivo.efectivo_nombre
            target['C10'] = turno.efectivo.efectivo_jerarquia
            target['D10'] = turno.efectivo.efectivo_legajo
            target['F10'] = turno.efectivo.efectivo_dni

            target['A6'] = adicional.nombre + " | " + str(adicional.numero)
            target['D6'] = str(adicional.comisaria) + " " + adicional.domicilio
            target['A10'] = adicional.categoria
            target['D8'] = adicional.localidad
            target['F8'] = adicional.localidad

            target['B10'] = request.session['export_month']

            if adicional.categoria == '1era' or adicional.categoria == '4ta':
                target['G29'].value = "=G28*(210-21)"
            elif adicional.categoria == '2da' or adicional.categoria == '3era':
                target['G29'].value = "=G28*(380-38)"
            else:
                target['G29'].value = "=G28*(380-38)"

            # corrige los bordes que se pierden en la copia de pestaña:

            for merged_cells in target.merged_cells.ranges:
                style = target.cell(merged_cells.min_row, merged_cells.min_col)._style
                for col in range(merged_cells.min_col, merged_cells.max_col + 1):
                    for row in range(merged_cells.min_row, merged_cells.max_row + 1):
                        target.cell(row, col)._style = style

    rend.remove(rend['Efectivo'])

    final_list.sort(key = lambda final_list: final_list[2])         # ordena por horarios

    # carga de turnos:

    j = 13
    for i in range (1, 17):
        for record in final_list:
            if record[1] == i:
                if rend[record[0]]['B'+str(j)].value is None:
                    rend[record[0]]['B'+str(j)].value = record[2]
                    rend[record[0]]['C'+str(j)].value = record[3]
                else:
                    rend[record[0]]['B'+str(j)].value = rend[record[0]]['B'+str(j)].value + ' y ' + record[2]
                    rend[record[0]]['C'+str(j)].value = rend[record[0]]['C'+str(j)].value + record[3]
                rend[record[0]]['G28'].value = rend[record[0]]['G28'].value + record[3]
        j += 1

    j = 13
    for i in range (17, 32):
        for record in final_list:
            if record[1] == i:
                if rend[record[0]]['E'+str(j)].value is None:
                    rend[record[0]]['E'+str(j)].value = record[2]
                    rend[record[0]]['G'+str(j)].value = record[3]
                else:
                    rend[record[0]]['E'+str(j)].value = rend[record[0]]['E'+str(j)].value + ' y ' + record[2]
                    rend[record[0]]['G'+str(j)].value = rend[record[0]]['G'+str(j)].value + record[3]
                rend[record[0]]['G28'].value = rend[record[0]]['G28'].value + record[3]
        j += 1

    # finaliza carga de turnos.

    # boleta general:

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    boletascio = rend['General']

    boletascio['E2'] = adicional.numero
    boletascio['I6'] = request.session['export_month']
    boletascio['E9'] = adicional.entidad_contratante
    boletascio['E10'] = adicional.localidad + ', ' + str(adicional.comisaria)
    boletascio['E11'] = adicional.domicilio + ', ' + adicional.localidad
    boletascio['A13'] = adicional.categoria
    boletascio['B13'] = adicional.cuenta_corriente
    boletascio['E13'].value = len(set(f[1] for f in final_list))

    if adicional.categoria == '1era' or adicional.categoria == '4ta':
        boletascio['G13'].value = "=F13*210"
    elif adicional.categoria == '2da' or adicional.categoria == '3era':
        boletascio['G13'].value = "=F13*380"
    else:
        boletascio['G13'].value = "=F13*360"

    pest = sorted(rend.sheetnames)
    cont = 17

    for p in sorted(rend.sheetnames):
        if p != 'General':
            orden = boletascio['A'+str(cont)]
            orden.value = cont - 16
            orden.fill = PatternFill("solid", fgColor="EFEFEF")
            orden.font = Font(name='Courier New', size=9, bold=True)
            boletascio['B'+str(cont)] = p
            ind = rend.get_sheet_by_name(name=p)
            boletascio['E'+str(cont)].value = ind['D10'].value
            boletascio['F'+str(cont)].value = ind['F10'].value
            boletascio['G'+str(cont)].value = ind['C10'].value
            boletascio['H'+str(cont)].value = ind['G28'].value

            if adicional.categoria == '1era' or adicional.categoria == '4ta':
                boletascio['I'+str(cont)].value = "=H"+str(cont)+"*(210-21)"
            elif adicional.categoria == '2da' or adicional.categoria == '3era':
                boletascio['I'+str(cont)].value = "=H"+str(cont)+"*(380-38)"
            else:
                boletascio['I'+str(cont)].value = "=H"+str(cont)+"*(360-36)"

            monto = boletascio['I'+str(cont)]
            monto.number_format = '$#,##0.00'

            boletascio.insert_rows(cont + 1)
            new_row = boletascio['B'+str(cont + 1) + ':' + 'H' + str(cont + 1)]
            boletascio.merge_cells('B'+str(cont + 1)+':'+'D'+str(cont + 1))
            boletascio.merge_cells('I'+str(cont + 1)+':'+'J'+str(cont + 1))

            for i, rowOfCellObjects in enumerate(boletascio['A'+str(cont + 1) + ':' + 'J' + str(cont + 1)]):
                for n, cellObj in enumerate(rowOfCellObjects):
                    cellObj.font = Font(name='Courier New', size=9)
                    cellObj.alignment = Alignment(horizontal="center", vertical="center")
                    cellObj.border = thin_border

            cont += 1

    boletascio.delete_rows(cont, amount=2)
    boletascio.unmerge_cells('B'+str(cont)+':'+'D'+str(cont))
    boletascio.merge_cells('I'+str(cont)+':'+'J'+str(cont))
    boletascio.merge_cells('I'+str(cont + 1)+':'+'J'+str(cont + 1))
    boletascio.merge_cells('I'+str(cont + 2)+':'+'J'+str(cont + 2))

    boletascio['G'+str(cont + 1)].value = "TOTAL HORAS"
    boletascio['H'+str(cont + 1)].value = "=SUM(H17:H{})".format(cont - 1)
    boletascio['F13'].value = "=SUM(H17:H{})".format(cont - 1)                   # total de horas en la parte superior
    boletascio['I'+str(cont + 1)].value = "=SUM(I17:I{})".format(cont - 1)
    boletascio['I'+str(cont + 2)].value = "TOTAL 90%"

    total_hs = boletascio['G'+str(cont + 1)]
    total_hs.font = Font(name='Courier New', size=9, bold=True)
    total_hs.border = thin_border
    total_hs.fill = PatternFill("solid", fgColor="EFEFEF")

    horas = boletascio['H'+str(cont + 1)]
    horas.font = Font(name='Courier New', size=9, bold=True)
    horas.border = thin_border
    horas.fill = PatternFill("solid", fgColor="EFEFEF")

    total = boletascio['I'+str(cont + 1)]
    total.number_format = '$#,##0.00'
    total.font = Font(name='Courier New', size=9, bold=True)
    total.fill = PatternFill("solid", fgColor="EFEFEF")

    for i, rowOfCellObjects in enumerate(boletascio['I'+str(cont + 1) + ':' + 'J' + str(cont + 1)]):
        for n, cellObj in enumerate(rowOfCellObjects):
            cellObj.border = thin_border

    total_90 = boletascio['I'+str(cont + 2)]
    total_90.font = Font(name='Courier New', size=9, bold=True)
    total_90.fill = PatternFill("solid", fgColor="EFEFEF")

    for i, rowOfCellObjects in enumerate(boletascio['I'+str(cont + 2) + ':' + 'J' + str(cont + 2)]):
        for n, cellObj in enumerate(rowOfCellObjects):
            cellObj.border = thin_border

    cut_row = cont + 4

    boletascio.merge_cells('A'+str(cut_row)+':'+'J'+str(cut_row))
    renglon_1 = boletascio['A'+str(cut_row)]
    renglon_1.value = 'Por la presente dejo constancia que se ha prestado de conformidad el Servicio de Policía Adicional durante el período comprendido entre el'
    renglon_1.alignment = Alignment(horizontal="left", vertical="center")

    boletascio.merge_cells('A'+str(cut_row + 1)+':'+'J'+str(cut_row + 1))
    renglon_2 = boletascio['A'+str(cut_row + 1)]
    renglon_2.value = str(min(turno.ingreso.day for turno in turnos)) + ' hasta el ' + str(max(turno.ingreso.day for turno in turnos)) + ' de ' + request.session['export_month'] + ' completando la cantidad de horas mencionadas. - - - - - - '
    renglon_2.alignment = Alignment(horizontal="left", vertical="center")

    cut_row = cut_row + 3

    boletascio['A'+str(cut_row)].value = "Lugar:"
    boletascio['C'+str(cut_row)].value = adicional.localidad
    boletascio['F'+str(cut_row)].value = "Fecha:"
    boletascio['G'+str(cut_row)].value = "_____"
    boletascio['H'+str(cut_row)].value = "de"
    boletascio['I'+str(cut_row)].value = "_________________"
    boletascio['J'+str(cut_row)].value = "del año"

    cut_row = cut_row + 10

    boletascio.merge_cells('A'+str(cut_row)+':'+'D'+str(cut_row))
    sello_usuario = boletascio['A'+str(cut_row)]
    sello_usuario.value = "________________________________________________"
    boletascio.merge_cells('A'+str(cut_row + 1)+':'+'D'+str(cut_row + 1))
    boletascio['A'+str(cut_row + 1)].value = "Sello Usuario"

    boletascio.merge_cells('G'+str(cut_row)+':'+'J'+str(cut_row))
    firma_aclaracion = boletascio['G'+str(cut_row)]
    firma_aclaracion.value = "________________________________________________"
    boletascio.merge_cells('G'+str(cut_row + 1)+':'+'J'+str(cut_row + 1))
    boletascio['G'+str(cut_row + 1)].value = "Firma y Aclaración"

    cut_row = cut_row + 4

    boletascio['F'+str(cut_row)].value = adicional.localidad
    boletascio['G'+str(cut_row)].value = "_____"
    boletascio['H'+str(cut_row)].value = "de"
    boletascio['I'+str(cut_row)].value = "_________________"
    boletascio['J'+str(cut_row)].value = "del año"

    cut_row = cut_row + 3

    boletascio.merge_cells('A'+str(cut_row)+':'+'J'+str(cut_row))
    renglon_1 = boletascio['A'+str(cut_row)]
    renglon_1.value = 'Se eleva la presente boleta de servicio a los fines de su liquidación, conforme con las certificaciones realizadas por el usuario y'
    renglon_1.alignment = Alignment(horizontal="left", vertical="center")

    boletascio.merge_cells('A'+str(cut_row + 1)+':'+'J'+str(cut_row + 1))
    renglon_2 = boletascio['A'+str(cut_row + 1)]
    renglon_2.value = 'la Dependencia Policial que interviene. - - - - - - '
    renglon_2.alignment = Alignment(horizontal="left", vertical="center")

    cut_row = cut_row + 10

    boletascio.merge_cells('A'+str(cut_row)+':'+'D'+str(cut_row))
    sello_usuario = boletascio['A'+str(cut_row)]
    sello_usuario.value = "________________________________________________"
    boletascio.merge_cells('A'+str(cut_row + 1)+':'+'D'+str(cut_row + 1))
    boletascio['A'+str(cut_row + 1)].value = "Sello Dependencia Policial"

    boletascio.merge_cells('G'+str(cut_row)+':'+'J'+str(cut_row))
    firma_aclaracion = boletascio['G'+str(cut_row)]
    firma_aclaracion.value = "________________________________________________"
    boletascio.merge_cells('G'+str(cut_row + 1)+':'+'J'+str(cut_row + 1))
    boletascio['G'+str(cut_row + 1)].value = "Firma y Aclaración"

    for sheet in rend.sheetnames:
        rend['{}'.format(sheet)].protection.sheet = True
        rend['{}'.format(sheet)].protection.password = 'macrigato'
        rend['{}'.format(sheet)].protection.enable()


    response = HttpResponse(content=save_virtual_workbook(rend), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Rendicion.xlsx'

    return response
