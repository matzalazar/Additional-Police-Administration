from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .forms import AgregarCores
from .models import Cores
from administracion.models import Efectivo, Comisaria
from polads.models import Turno
from users.models import Profile
from django.http import JsonResponse
from django.core import serializers
from datetime import datetime, timedelta
from .filters import CoresFilter
from django.db.models import F, Sum, ExpressionWrapper, fields, Case, When, Value, CharField
from mensajes.models import Mensaje
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Create your views here.

# cores create record

@login_required
def agregar_cores(request):
    cores = Cores.objects.filter(dependencia=request.user.profile.administrador_de).order_by('-pk')[:10]
    dependencia = Comisaria.objects.get(dependencia=request.user.profile.administrador_de)
    errors = ''

    form = AgregarCores()

    if request.method == 'POST':
        form = AgregarCores(request.POST)

        if form.is_valid():

            instance = form.save(commit=False)

            # validación de turnos:

            all_turnos = Turno.objects.filter(efectivo=instance.efectivo)
            all_cores = Cores.objects.filter(efectivo=instance.efectivo)

            cant_cores = Cores.objects.filter(efectivo=instance.efectivo, ingreso__month=instance.ingreso.month).aggregate(
            horas=Sum(ExpressionWrapper(F('egreso') - F('ingreso'), output_field=fields.DurationField())))

            cant_polad = Turno.objects.filter(efectivo=instance.efectivo, ingreso__month=instance.ingreso.month).aggregate(
            horas=Sum(ExpressionWrapper(F('egreso') - F('ingreso'), output_field=fields.DurationField())))

            try:
                cant_cores = int(cant_cores.get('horas').total_seconds()/60/60)
            except:
                cant_cores = 0

            try:
                cant_polad = int(cant_polad.get('horas').total_seconds()/60/60)
            except:
                cant_polad = 0

            horas_mensuales = cant_cores + cant_polad

            es_valido = True

            for each in all_turnos:
                if instance.ingreso < each.egreso and each.ingreso < instance.egreso:
                        errors = 'Turno inválido para el efectivo ' + str(instance.efectivo) + ". Comunicate con " + str(each.polad.encargado) + "."
                        try:
                            administrador = Profile.objects.get(administrador_de=each.polad.comisaria)
                            mensaje = Mensaje(remitente=request.user, destinatario=administrador.user, mensaje=f'Mensaje automático. El usuario {request.user} ha intentado cargar un turno inválido de Cores para el efectivo {instance.efectivo} en Comisaría {request.user.profile.administrador_de}  el día {instance.ingreso.day}/{instance.ingreso.month} a partir de las {instance.ingreso.hour}.')
                            mensaje.save()
                        except Profile.DoesNotExist:
                            administrador = None
                        es_valido = False

            for each in all_cores:
                if instance.ingreso < each.egreso and each.ingreso < instance.egreso:
                        responsable = Profile.objects.get(administrador_de=each.dependencia)
                        errors = 'Turno inválido para el efectivo ' + str(instance.efectivo) + ". Comunicate con " + str(responsable.user) + "."
                        mensaje = Mensaje(remitente=request.user, destinatario=responsable.user, mensaje=f'Mensaje automático. El usuario {request.user} ha intentado cargar un turno inválido de Cores para el efectivo {instance.efectivo} en Comisaría {request.user.profile.administrador_de} el día {instance.ingreso.day}/{instance.ingreso.month} a partir de las {instance.ingreso.hour}.')
                        mensaje.save()
                        es_valido = False

            cant_instance = int((instance.egreso - instance.ingreso).seconds/3600)

            if instance.ingreso > instance.egreso:
                errors = "El ingreso no puede ser posterior al egreso."
                es_valido = False

            if cant_instance + cant_cores > 80:
                errors = "El efectivo ha superado el límite de horas para hacer Cores."
                es_valido = False

            # procesar turnos y horarios:

            if es_valido == True:

                if instance.ingreso.day == instance.egreso.day:
                    instance_valid = Cores(ingreso=instance.ingreso, egreso=instance.egreso, efectivo=instance.efectivo, tipo=instance.tipo, dependencia=dependencia)
                    instance_valid.save()

                elif instance.ingreso.day != instance.egreso.day:

                    if instance.egreso.month - instance.ingreso.month == 1 or instance.egreso.year - instance.ingreso.year == 1:
                        instance_valid_1 = Cores(ingreso=instance.ingreso, egreso=instance.egreso.replace(hour=00), efectivo=instance.efectivo, tipo=instance.tipo, dependencia=dependencia)
                        instance_valid_1.save()
                        instance_valid_2 = Cores(ingreso=instance.egreso.replace(hour=00), egreso=instance.egreso, efectivo=instance.efectivo, tipo=instance.tipo, dependencia=dependencia)
                        if instance_valid_2.egreso.hour == 00: pass
                        else: instance_valid_2.save()

                    if instance.egreso.day - instance.ingreso.day == 1:
                       instance_valid_1 = Cores(ingreso=instance.ingreso, egreso=instance.egreso.replace(hour=00), efectivo=instance.efectivo, tipo=instance.tipo, dependencia=dependencia)
                       instance_valid_1.save()
                       instance_valid_2 = Cores(ingreso=instance.egreso.replace(hour=00), egreso=instance.egreso, efectivo=instance.efectivo, tipo=instance.tipo, dependencia=dependencia)
                       if instance_valid_2.egreso.hour == 00: pass
                       else: instance_valid_2.save()

                    if instance.egreso.day - instance.ingreso.day > 1:
                        delta = instance.egreso - instance.ingreso
                        instance_valid_1 = Cores(ingreso=instance.ingreso, egreso=instance.ingreso.replace(day=instance.ingreso.day + 1, hour=00), efectivo=instance.efectivo, tipo=instance.tipo, dependencia=dependencia)
                        instance_valid_1.save()
                        for i in range(delta.days):
                            day = instance.ingreso + timedelta(days=i)
                            instance_valid_2 = Cores(ingreso=day.replace(day=day.day + 1, hour=00), egreso=day.replace(day=day.day + 2, hour=00), efectivo=instance.efectivo, tipo=instance.tipo, dependencia=dependencia)
                            if instance_valid_2.ingreso.day == instance.egreso.day: pass
                            else: instance_valid_2.save()
                        instance_valid_3 = Cores(ingreso=instance.egreso.replace(hour=00), egreso=instance.egreso, efectivo=instance.efectivo, tipo=instance.tipo, dependencia=dependencia)
                        instance_valid_3.save()

        else:
            print(form.errors.items())

    context = {'cores': cores, 'form': form, 'errors': errors}
    return render(request, 'cores/agregar_cores.html', context)

# cores delete record

@login_required
def deletecores(request, field):
    previous_page = request.META['HTTP_REFERER']
    corestodelete = Cores.objects.get(id=field)
    corestodelete.delete()
    return redirect(previous_page)

# view cores by filter

@login_required
def ver_cores(request):
    cores = Cores.objects.filter(dependencia=request.user.profile.administrador_de).order_by('-ingreso')
    myFilter = CoresFilter(request.GET, queryset=cores)
    cores = myFilter.qs
    context = {'cores': cores, 'myFilter': myFilter}
    return render(request, 'cores/ver_cores.html', context)

# cupo cores

@login_required
def cupo_cores(request):
    today = datetime.now()

    cores = Cores.objects.filter(dependencia=request.user.profile.administrador_de, ingreso__month=today.month).annotate(
    custom_order_1=Case(
    When(efectivo__efectivo_jerarquia='CRIO', then=Value(1)),
    When(efectivo__efectivo_jerarquia='SUBCRIO', then=Value(2)),
    When(efectivo__efectivo_jerarquia='PPAL', then=Value(3)),
    When(efectivo__efectivo_jerarquia='OI', then=Value(4)),
    When(efectivo__efectivo_jerarquia='OSI', then=Value(5)),
    When(efectivo__efectivo_jerarquia='OA', then=Value(6)),
    When(efectivo__efectivo_jerarquia='OSA', then=Value(7)),
    When(efectivo__efectivo_jerarquia='MAYOR', then=Value(8)),
    When(efectivo__efectivo_jerarquia='CAP', then=Value(9)),
    When(efectivo__efectivo_jerarquia='TTE1ERO', then=Value(10)),
    When(efectivo__efectivo_jerarquia='TTE', then=Value(11)),
    When(efectivo__efectivo_jerarquia='SUBTTE', then=Value(12)),
    When(efectivo__efectivo_jerarquia='SGTO', then=Value(13)),
    When(efectivo__efectivo_jerarquia='OFL', then=Value(14)),
    output_field=CharField())).order_by('custom_order_1', 'efectivo__efectivo_legajo').values(
    'efectivo__efectivo_item', 'efectivo__efectivo_nombre', 'efectivo__efectivo_legajo', 'efectivo__efectivo_jerarquia', 'tipo').annotate(
    horas=Sum(ExpressionWrapper(F('egreso') - F('ingreso'), output_field=fields.DurationField())))


    # OPERATIVAS

    cupo_operativas = Cores.objects.filter(dependencia=request.user.profile.administrador_de, ingreso__month=today.month, tipo='OP').aggregate(
    horas=Sum(ExpressionWrapper(F('egreso') - F('ingreso'), output_field=fields.DurationField())))

    try:
        cupo_operativas = int(cupo_operativas.get('horas').total_seconds()/60/60)
    except:
        cupo_operativas = 0

    # OPERATIVAS CHOFERES

    cupo_operativas_choferes = Cores.objects.filter(dependencia=request.user.profile.administrador_de, ingreso__month=today.month, tipo='OP CHOF').aggregate(
    horas=Sum(ExpressionWrapper(F('egreso') - F('ingreso'), output_field=fields.DurationField())))

    try:
        cupo_operativas_choferes = int(cupo_operativas_choferes.get('horas').total_seconds()/60/60)
    except:
        cupo_operativas_choferes = 0

    # NO OPERATIVAS

    cupo_no_operativas = Cores.objects.filter(dependencia=request.user.profile.administrador_de, ingreso__month=today.month, tipo='NO OP').aggregate(
    horas=Sum(ExpressionWrapper(F('egreso') - F('ingreso'), output_field=fields.DurationField())))

    try:
        cupo_no_operativas = int(cupo_no_operativas.get('horas').total_seconds()/60/60)
    except:
        cupo_no_operativas = 0

    context = {'cores': cores, 'cupo_operativas': cupo_operativas, 'cupo_operativas_choferes': cupo_operativas_choferes, 'cupo_no_operativas': cupo_no_operativas, 'today': today}
    return render(request, 'cores/cupo_cores.html', context)

# rendición de cores:

@login_required
def rendir_cores(request):

    today = datetime.now()
    mes = today.month
    show = ''

    if mes == 1: show = "Enero de {}".format(today.year)
    elif mes == 2: show = "Febrero de {}".format(today.year)
    elif mes == 3: show = "Marzo de {}".format(today.year)
    elif mes == 4: show = "Abril de {}".format(today.year)
    elif mes == 5: show = "Mayo de {}".format(today.year)
    elif mes == 6: show = "Junio de {}".format(today.year)
    elif mes == 7: show = "Julio de {}".format(today.year)
    elif mes == 8: show = "Agosto de {}".format(today.year)
    elif mes == 9: show = "Septiembre de {}".format(today.year)
    elif mes == 10: show = "Octubre de {}".format(today.year)
    elif mes == 11: show = "Noviembre de {}".format(today.year)
    elif mes == 12: show = "Diciembre de {}".format(today.year)

    cores_op = Cores.objects.filter(dependencia=request.user.profile.administrador_de, ingreso__month=today.month, tipo='OP').annotate(
    custom_order=Case(
    When(efectivo__efectivo_jerarquia='CRIO', then=Value(14)),
    When(efectivo__efectivo_jerarquia='SUBCRIO', then=Value(13)),
    When(efectivo__efectivo_jerarquia='PPAL', then=Value(12)),
    When(efectivo__efectivo_jerarquia='OI', then=Value(11)),
    When(efectivo__efectivo_jerarquia='OSI', then=Value(10)),
    When(efectivo__efectivo_jerarquia='OA', then=Value(9)),
    When(efectivo__efectivo_jerarquia='OSA', then=Value(8)),
    When(efectivo__efectivo_jerarquia='MAYOR', then=Value(7)),
    When(efectivo__efectivo_jerarquia='CAP', then=Value(6)),
    When(efectivo__efectivo_jerarquia='TTE1ERO', then=Value(5)),
    When(efectivo__efectivo_jerarquia='TTE', then=Value(4)),
    When(efectivo__efectivo_jerarquia='SUBTTE', then=Value(3)),
    When(efectivo__efectivo_jerarquia='SGTO', then=Value(2)),
    When(efectivo__efectivo_jerarquia='OFL', then=Value(1)),
    output_field=CharField())).order_by('-custom_order', 'efectivo__efectivo_legajo').values(
    'efectivo__efectivo_item', 'efectivo__efectivo_nombre', 'efectivo__efectivo_legajo', 'efectivo__efectivo_jerarquia', 'tipo').annotate(
    horas=Sum(ExpressionWrapper(F('egreso') - F('ingreso'), output_field=fields.DurationField())))

    cores_op_chof = Cores.objects.filter(dependencia=request.user.profile.administrador_de, ingreso__month=today.month, tipo='OP CHOF').annotate(
    custom_order=Case(
    When(efectivo__efectivo_jerarquia='CRIO', then=Value(14)),
    When(efectivo__efectivo_jerarquia='SUBCRIO', then=Value(13)),
    When(efectivo__efectivo_jerarquia='PPAL', then=Value(12)),
    When(efectivo__efectivo_jerarquia='OI', then=Value(11)),
    When(efectivo__efectivo_jerarquia='OSI', then=Value(10)),
    When(efectivo__efectivo_jerarquia='OA', then=Value(9)),
    When(efectivo__efectivo_jerarquia='OSA', then=Value(8)),
    When(efectivo__efectivo_jerarquia='MAYOR', then=Value(7)),
    When(efectivo__efectivo_jerarquia='CAP', then=Value(6)),
    When(efectivo__efectivo_jerarquia='TTE1ERO', then=Value(5)),
    When(efectivo__efectivo_jerarquia='TTE', then=Value(4)),
    When(efectivo__efectivo_jerarquia='SUBTTE', then=Value(3)),
    When(efectivo__efectivo_jerarquia='SGTO', then=Value(2)),
    When(efectivo__efectivo_jerarquia='OFL', then=Value(1)),
    output_field=CharField())).order_by('-custom_order', 'efectivo__efectivo_legajo').values(
    'efectivo__efectivo_item', 'efectivo__efectivo_nombre', 'efectivo__efectivo_legajo', 'efectivo__efectivo_jerarquia', 'tipo').annotate(
    horas=Sum(ExpressionWrapper(F('egreso') - F('ingreso'), output_field=fields.DurationField())))

    cores_no_op = Cores.objects.filter(dependencia=request.user.profile.administrador_de, ingreso__month=today.month, tipo='NO OP').annotate(
    custom_order=Case(
    When(efectivo__efectivo_jerarquia='CRIO', then=Value(14)),
    When(efectivo__efectivo_jerarquia='SUBCRIO', then=Value(13)),
    When(efectivo__efectivo_jerarquia='PPAL', then=Value(12)),
    When(efectivo__efectivo_jerarquia='OI', then=Value(11)),
    When(efectivo__efectivo_jerarquia='OSI', then=Value(10)),
    When(efectivo__efectivo_jerarquia='OA', then=Value(9)),
    When(efectivo__efectivo_jerarquia='OSA', then=Value(8)),
    When(efectivo__efectivo_jerarquia='MAYOR', then=Value(7)),
    When(efectivo__efectivo_jerarquia='CAP', then=Value(6)),
    When(efectivo__efectivo_jerarquia='TTE1ERO', then=Value(5)),
    When(efectivo__efectivo_jerarquia='TTE', then=Value(4)),
    When(efectivo__efectivo_jerarquia='SUBTTE', then=Value(3)),
    When(efectivo__efectivo_jerarquia='SGTO', then=Value(2)),
    When(efectivo__efectivo_jerarquia='OFL', then=Value(1)),
    output_field=CharField())).order_by('-custom_order', 'efectivo__efectivo_legajo').values(
    'efectivo__efectivo_item', 'efectivo__efectivo_nombre', 'efectivo__efectivo_legajo', 'efectivo__efectivo_jerarquia', 'tipo').annotate(
    horas=Sum(ExpressionWrapper(F('egreso') - F('ingreso'), output_field=fields.DurationField())))

    rend = load_workbook('/administracion/files/Cores.xlsx')

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # rendición cores operativas

    op = rend['Op']
    op['A4'].value = str(request.user.profile.administrador_de)
    op['A6'].value = "Mes de " + show

    cont = 9

    for record in cores_op:

        op.row_dimensions[cont].height = 23.05
        orden = op['A'+str(cont)]
        orden.value = cont - 8
        orden.font = Font(name='Courier New', size=8, bold=True)
        op['B'+str(cont)].value = record.get('efectivo__efectivo_jerarquia')
        op['C'+str(cont)].value = record.get('efectivo__efectivo_legajo')
        op['D'+str(cont)].value = record.get('efectivo__efectivo_item')
        op['E'+str(cont)].value = record.get('efectivo__efectivo_nombre')
        op['G'+str(cont)].value = int(record.get('horas').total_seconds()/60/60)

        op.merge_cells('E'+str(cont + 1)+':'+'F'+str(cont + 1))

        for i, rowOfCellObjects in enumerate(op['A'+str(cont) + ':' + 'H' + str(cont)]):
            for n, cellObj in enumerate(rowOfCellObjects):
                cellObj.font = Font(name='Courier New', size=8)
                cellObj.alignment = Alignment(horizontal="center", vertical="center")
                cellObj.border = thin_border

        cont += 1

    new_row = cont + 2
    op.merge_cells('A'+str(new_row)+':'+'F'+str(new_row))
    op.row_dimensions[new_row].height = 23.05

    total_hs = op['A'+str(new_row)]
    total_hs.value = "TOTAL: "
    total_hs.font = Font(name='Courier New', size=8, bold=True)
    total_hs.border = thin_border

    horas = op['G'+str(new_row)]
    horas.value = "=SUM(G9:H{})".format(new_row - 1)
    horas.font = Font(name='Courier New', size=8, bold=True)
    horas.border = thin_border

    # rendición cores de chofer

    op_chof = rend['Op Chof']
    op_chof['A4'].value = str(request.user.profile.administrador_de)
    op_chof['A6'].value = "Mes de " + show

    cont = 9

    for record in cores_op_chof:

        op_chof.row_dimensions[cont].height = 23.05
        orden = op_chof['A'+str(cont)]
        orden.value = cont - 8
        orden.font = Font(name='Courier New', size=8, bold=True)
        op_chof['B'+str(cont)].value = record.get('efectivo__efectivo_jerarquia')
        op_chof['C'+str(cont)].value = record.get('efectivo__efectivo_legajo')
        op_chof['D'+str(cont)].value = record.get('efectivo__efectivo_item')
        op_chof['E'+str(cont)].value = record.get('efectivo__efectivo_nombre')
        op_chof['G'+str(cont)].value = int(record.get('horas').total_seconds()/60/60)

        op_chof.merge_cells('E'+str(cont + 1)+':'+'F'+str(cont + 1))

        for i, rowOfCellObjects in enumerate(op_chof['A'+str(cont) + ':' + 'H' + str(cont)]):
            for n, cellObj in enumerate(rowOfCellObjects):
                cellObj.font = Font(name='Courier New', size=8)
                cellObj.alignment = Alignment(horizontal="center", vertical="center")
                cellObj.border = thin_border

        cont += 1

    new_row = cont + 2
    op_chof.merge_cells('A'+str(new_row)+':'+'F'+str(new_row))
    op_chof.row_dimensions[new_row].height = 23.05

    total_hs = op_chof['A'+str(new_row)]
    total_hs.value = "TOTAL: "
    total_hs.font = Font(name='Courier New', size=8, bold=True)
    total_hs.border = thin_border

    horas = op_chof['G'+str(new_row)]
    horas.value = "=SUM(G9:H{})".format(new_row - 1)
    horas.font = Font(name='Courier New', size=8, bold=True)
    horas.border = thin_border

    # rendición cores no operativas

    no_op = rend['No Op']
    no_op['A4'].value = str(request.user.profile.administrador_de)
    no_op['A6'].value = "Mes de " + show

    cont = 9

    for record in cores_no_op:

        no_op.row_dimensions[cont].height = 23.05
        orden = no_op['A'+str(cont)]
        orden.value = cont - 8
        orden.font = Font(name='Courier New', size=8, bold=True)
        no_op['B'+str(cont)].value = record.get('efectivo__efectivo_jerarquia')
        no_op['C'+str(cont)].value = record.get('efectivo__efectivo_legajo')
        no_op['D'+str(cont)].value = record.get('efectivo__efectivo_item')
        no_op['E'+str(cont)].value = record.get('efectivo__efectivo_nombre')
        no_op['G'+str(cont)].value = int(record.get('horas').total_seconds()/60/60)

        no_op.merge_cells('E'+str(cont + 1)+':'+'F'+str(cont + 1))

        for i, rowOfCellObjects in enumerate(no_op['A'+str(cont) + ':' + 'H' + str(cont)]):
            for n, cellObj in enumerate(rowOfCellObjects):
                cellObj.font = Font(name='Courier New', size=8)
                cellObj.alignment = Alignment(horizontal="center", vertical="center")
                cellObj.border = thin_border

        cont += 1

    new_row = cont + 2
    no_op.merge_cells('A'+str(new_row)+':'+'F'+str(new_row))
    no_op.row_dimensions[new_row].height = 23.05

    total_hs = no_op['A'+str(new_row)]
    total_hs.value = "TOTAL: "
    total_hs.font = Font(name='Courier New', size=8, bold=True)
    total_hs.border = thin_border

    horas = no_op['G'+str(new_row)]
    horas.value = "=SUM(G9:H{})".format(new_row - 1)
    horas.font = Font(name='Courier New', size=8, bold=True)
    horas.border = thin_border

    if not cores_op:
        rend.remove(rend['Op'])
    elif not cores_op_chof:
        rend.remove(rend['Op Chof'])
    elif not cores_no_op:
        rend.remove(rend['No Op'])

    response = HttpResponse(content=save_virtual_workbook(rend), content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Cores.xlsx'

    return response
